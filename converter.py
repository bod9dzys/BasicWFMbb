# -*- coding: utf-8 -*-
import csv
import re
from datetime import datetime, time, timedelta

import openpyxl
from openpyxl import Workbook
# Видалено unidecode, оскільки імена залишаються українською
# from unidecode import unidecode

# =========================
# 1) НАЛАШТУВАННЯ
# =========================

# Вхідний Excel з графіком
INPUT_FILE = "november.xlsx" # Припустимо, новий файл називається так

# Назва аркуша (None = перший активний)
SHEET_NAME = None

# Формат вихідного файлу: 'xlsx' або 'csv'
OUTPUT_FORMAT = "xlsx"   # 'xlsx' або 'csv'

# Ім'я вихідного файлу (без розширення, ми додамо самі)
OUTPUT_BASENAME = "import_for_wfm_ukr_names_tl" # Змінено ім'я для ясності

# Для CSV: щоб Excel в українській локалі не склеював колонки
CSV_DELIMITER = ";"
CSV_ENCODING = "utf-8" # utf-8-sig може бути краще для Excel з кирилицею

# Мапа статусів
STATUS_MAP = {
    "Відпустка": "vacation",
    "OFF": "day_off",
    "вихідний": "day_off",
    "Лікарняний": "sick",
    # додай інші за потреби
}

# Спеціальні текстові мітки для статусу/активності
SPECIAL_ACTIVITY_MAP = {}

# Назви колонок - ДОДАНО 'team_lead'
HEADERS = ["id", "agent", "team_lead", "start", "end", "direction", "status", "activity", "comment"]


# =========================
# 2) ГЕНЕРАЦІЯ USERNAME (Залишаємо для потенційного використання, але не для колонки 'agent')
# =========================
def generate_username(full_name: str) -> str:
    """
    Генерує ЛАТИНСЬКИЙ username з повного імені для Django User model.
    Приклад: "Ступак Максим" -> "mstupak"
    """
    # Використовуємо unidecode тут, ЛОКАЛЬНО для генерації логіну
    from unidecode import unidecode # Імпортуємо тут, щоб не було глобальної залежності
    latin_name = unidecode(full_name)
    parts = latin_name.split()
    if len(parts) >= 2:
        first_name = parts[0]
        last_name = parts[1]
        username_raw = f"{first_name[0]}{last_name}"
    elif len(parts) == 1:
        username_raw = parts[0]
    else:
        return f"user_{int(datetime.now().timestamp())}"

    username_clean = re.sub(r"[^a-zA-Z0-9]", "", username_raw).lower()
    return username_clean


# =========================
# 3) ДОПОМОЖНІ РЕЧІ ДЛЯ ВИВОДУ
# =========================
class OutputWriter:
    """
    Єдиний інтерфейс на два вихідних формати.
    Викликати .write_row(dict) для запису рядка
    і .close() після завершення.
    """

    def __init__(self, format_: str, basename: str):
        self.format_ = format_.lower().strip()
        self.rows_written = 0
        self.filepath = None

        if self.format_ == "xlsx":
            self.filepath = f"{basename}.xlsx"
            self.wb: Workbook = Workbook()
            self.ws = self.wb.active
            self.ws.title = "export"
            self.ws.append(HEADERS) # <-- Використовує оновлені HEADERS
            self.is_csv = False

        elif self.format_ == "csv":
            self.filepath = f"{basename}.csv"
            try:
                # Використовуємо utf-8-sig для кращої сумісності з Excel та кирилицею
                self._csv_file = open(self.filepath, mode="w", encoding="utf-8-sig", newline="")
                self._writer = csv.writer(self._csv_file, delimiter=CSV_DELIMITER)
                self._writer.writerow(HEADERS) # <-- Використовує оновлені HEADERS
                self.is_csv = True
            except Exception as e:
                print(f"!!! ПОМИЛКА: Не вдалося відкрити CSV файл '{self.filepath}' для запису. {e}")
                self.filepath = None
                raise
        else:
            raise ValueError("OUTPUT_FORMAT має бути 'xlsx' або 'csv'.")

    @staticmethod
    def _fmt_dt(dt: datetime) -> str:
        return dt.strftime("%Y-%m-%d %H:%M")

    def write_row(self, row: dict):
        if self.filepath is None:
             return

        # Формуємо рядок згідно з оновленими HEADERS
        values = [
            row.get("id", ""),
            row.get("agent", ""), # <-- Тут буде українське ім'я
            row.get("team_lead", ""), # <-- Нове поле
            self._fmt_dt(row["start"]) if isinstance(row.get("start"), datetime) else row.get("start", ""),
            self._fmt_dt(row["end"]) if isinstance(row.get("end"), datetime) else row.get("end", ""),
            row.get("direction", ""),
            row.get("status", ""),
            row.get("activity", ""),
            row.get("comment", ""),
        ]
        if self.is_csv:
            self._writer.writerow(values)
        else:
            self.ws.append(values)
        self.rows_written += 1

    def close(self):
        if self.filepath is None:
            return

        if self.is_csv:
            if hasattr(self, '_csv_file') and self._csv_file:
                self._csv_file.close()
        else:
            for col_idx, header in enumerate(HEADERS, start=1):
                try:
                    column_letter = openpyxl.utils.get_column_letter(col_idx)
                    max_len = len(str(header))
                    # Обережно використовуємо str() для всіх значень
                    for cell in self.ws[column_letter]:
                         if cell.row > 1 and cell.value is not None:
                             try:
                                 cell_str = str(cell.value)
                                 # Обмежуємо довжину, щоб уникнути занадто широких стовпців
                                 max_len = max(max_len, len(cell_str))
                             except Exception:
                                 pass # Ігноруємо помилки перетворення в рядок для довжини
                    # Обмежуємо максимальну ширину
                    self.ws.column_dimensions[column_letter].width = min(max_len + 3, 50) # +3 для невеликого запасу
                except Exception as e:
                    print(f"Помилка при налаштуванні ширини колонки {col_idx}: {e}")
            try:
                self.wb.save(self.filepath)
            except Exception as e:
                 print(f"!!! ПОМИЛКА: Не вдалося зберегти XLSX файл '{self.filepath}'. Перевірте, чи він не відкритий іншою програмою. {e}")
                 self.filepath = None


# =========================
# 4) ОСНОВНА ЛОГІКА
# =========================
def convert_schedule_xlsx(input_path, output_basename, sheet_name=None, output_format="xlsx"):
    print(f"Відкриваю XLSX файл: {input_path}...")

    try:
        workbook = openpyxl.load_workbook(input_path, data_only=True)
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        print(f"Читаю аркуш: '{sheet.title}'")
    except FileNotFoundError:
        print(f"!!! ПОМИЛКА: Файл не знайдено за шляхом '{input_path}'")
        return
    except Exception as e:
        print(f"!!! ПОМИЛКА: Не можу відкрити файл '{input_path}'. {e}")
        return

    try:
        writer = OutputWriter(output_format, output_basename)
    except Exception:
        return

    # Залишаємо кеш для username, хоча він не йде прямо в вихідний файл
    generated_usernames_cache = {}

    header_row = sheet[1]
    dates = []
    # Коректно читаємо дати з D1 і далі
    for col_idx, cell in enumerate(header_row, start=1):
         if col_idx <= 3:
            continue
         date_val = cell.value
         parsed_date = None
         if isinstance(date_val, datetime):
             parsed_date = date_val.date()
         elif isinstance(date_val, (int, float)) and date_val > 0:
             try:
                 if 1 < date_val < 2958466:
                     excel_epoch = datetime(1899, 12, 30)
                     delta = timedelta(days=date_val)
                     parsed_datetime = excel_epoch + delta
                     parsed_date = parsed_datetime.date()
                 else:
                    print(f" [Warning] Значення '{date_val}' в колонці {col_idx} ({cell.coordinate}) схоже на число, але поза межами типових дат Excel. Ігнорується.")
             except Exception as e:
                  print(f" [Warning] Не вдалося конвертувати число '{date_val}' з Excel дати в колонці {col_idx} ({cell.coordinate}). {e}")
         elif date_val:
             date_str = str(date_val).split(" ")[0]
             for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%y", "%d.%m.%y"):
                 try:
                     parsed_date = datetime.strptime(date_str, fmt).date()
                     break
                 except (ValueError, TypeError):
                     continue
             if not parsed_date:
                 print(f" [Warning] Не вдалося розпізнати дату '{date_val}' в колонці {col_idx} ({cell.coordinate}) заголовку.")
         dates.append(parsed_date)

    valid_dates_count = len([d for d in dates if d])
    if valid_dates_count == 0:
         print("!!! ПОМИЛКА: Не знайдено жодної коректної дати в заголовку (рядок 1, починаючи з колонки D). Перевірте формат дат (напр., DD.MM.YYYY).")
         writer.close()
         return
    else:
        print(f"Знайдено {valid_dates_count} коректних дат в заголовку.")

    processed_shifts = 0

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        # Читаємо дані з нових колонок
        team_lead_name_cell = row[0].value # Колонка A - TeamLead
        direction_cell = row[1].value      # Колонка B - Direction
        agent_full_name_cell = row[2].value # Колонка C - Agent

        if not agent_full_name_cell:
            continue

        agent_full_name = str(agent_full_name_cell).strip()
        team_lead_name = str(team_lead_name_cell).strip() if team_lead_name_cell else "" # Зберігаємо ім'я ТЛ
        agent_direction = str(direction_cell).strip().lower() if direction_cell else "calls"

        # Генеруємо username (для можливого використання в Django), але не використовуємо його в колонці 'agent'
        if agent_full_name not in generated_usernames_cache:
            agent_username_generated = generate_username(agent_full_name) # Генеруємо латинський логін
            generated_usernames_cache[agent_full_name] = agent_username_generated # Зберігаємо його в кеш за повним ім'ям

        shifts_cells = row[3:]

        for date_obj, shift_cell in zip(dates, shifts_cells):
            if not date_obj or shift_cell.value is None:
                continue

            shift_raw = str(shift_cell.value).strip()
            if not shift_raw:
                continue

            try:
                # В `out` тепер передаємо УКРАЇНСЬКЕ ім'я агента та ТЛ
                out = {
                    "id": "",
                    "agent": agent_full_name, # <-- Українське ім'я
                    "team_lead": team_lead_name, # <-- Нове поле з ім'ям ТЛ
                    "direction": agent_direction,
                    "activity": "",
                    "comment": "",
                    "status": "work"
                }
                start_dt = None
                end_dt = None
                special_status = None
                special_activity_text = ""

                if shift_raw in STATUS_MAP:
                    out["status"] = STATUS_MAP[shift_raw]
                    start_dt = datetime.combine(date_obj, time(0, 0))
                    end_dt = start_dt + timedelta(days=1)
                elif ',' in shift_raw:
                    time_part, activity_part = shift_raw.split(',', 1)
                    time_part = time_part.strip()
                    activity_part = activity_part.strip()
                    activity_key = activity_part.lower()

                    if activity_key in SPECIAL_ACTIVITY_MAP:
                        special_status, special_activity_text = SPECIAL_ACTIVITY_MAP[activity_key]
                        if special_status:
                            out["status"] = special_status
                        out["activity"] = special_activity_text
                    else:
                        out["activity"] = activity_part

                    time_parts = time_part.split('-')
                    if len(time_parts) == 2:
                        start_str, end_str = map(str.strip, time_parts)
                        try:
                             h_start, m_start = map(int, start_str.split(':'))
                             start_dt = datetime.combine(date_obj, time(h_start, m_start))

                             if end_str == "24:00":
                                 end_dt = datetime.combine(date_obj + timedelta(days=1), time(0, 0))
                             else:
                                 h_end, m_end = map(int, end_str.split(':'))
                                 end_dt = datetime.combine(date_obj, time(h_end, m_end))

                             if end_dt <= start_dt:
                                 end_dt += timedelta(days=1)
                        except ValueError:
                             raise ValueError(f"Невірний формат часу '{time_part}' у записі з комою.")
                    else:
                        raise ValueError(f"Невірний формат часу '{time_part}' у записі з комою.")
                else:
                    parts = shift_raw.split("-")
                    if len(parts) == 2:
                        start_str, end_str = map(str.strip, parts)
                        try:
                             h_start, m_start = map(int, start_str.split(':'))
                             start_dt = datetime.combine(date_obj, time(h_start, m_start))

                             if end_str == "24:00":
                                 end_dt = datetime.combine(date_obj + timedelta(days=1), time(0, 0))
                             else:
                                 h_end, m_end = map(int, end_str.split(':'))
                                 end_dt = datetime.combine(date_obj, time(h_end, m_end))

                             if end_dt <= start_dt:
                                 end_dt += timedelta(days=1)
                        except ValueError:
                             raise ValueError(f"Невірний формат часу '{shift_raw}'. Очікується 'HH:MM-HH:MM'.")
                    else:
                        raise ValueError(f"Невідомий формат значення '{shift_raw}'. Очікується статус, 'HH:MM-HH:MM' або 'HH:MM-HH:MM, Текст'.")

                if start_dt and end_dt:
                    out["start"] = start_dt
                    out["end"] = end_dt
                    writer.write_row(out)
                    processed_shifts += 1
                elif out["status"] != "work": # Запис статусів на весь день
                     # Для статусів на весь день теж треба встановити start/end
                     out["start"] = datetime.combine(date_obj, time(0, 0))
                     out["end"] = out["start"] + timedelta(days=1)
                     writer.write_row(out)
                     processed_shifts += 1

            except Exception as e:
                print(
                    f"!!! ПОМИЛКА в рядку {row_idx} (Excel рядок): Не вдалося обробити клітинку {shift_cell.coordinate}: "
                    f"Агент='{agent_full_name}', ТЛ='{team_lead_name}', Дата='{date_obj}', Значення='{shift_raw}'"
                )
                print(f"     Текст помилки: {e}")

    writer.close()

    print("-" * 30)
    print(f"Готово! Оброблено {processed_shifts} змін.")
    if writer.filepath:
        print(f"Результат збережено у файл: {writer.filepath}")
    else:
        print("!!! Не вдалося зберегти вихідний файл через помилку.")


# =========================
# 5) ЗАПУСК
# =========================
if __name__ == "__main__":
    convert_schedule_xlsx(
        INPUT_FILE,
        OUTPUT_BASENAME,
        sheet_name=SHEET_NAME,
        output_format=OUTPUT_FORMAT,
    )